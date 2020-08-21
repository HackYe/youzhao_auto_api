# -*- coding: utf-8 -*-

"""
@Author: YuanYe
@time: 2020/7/9 15:04
"""

import openpyxl
from Common import dir_config
import re


class HandExcel:
    def load_excel(self):
        '''
        加载excel
        '''
        open_excel = openpyxl.load_workbook(dir_config.excel_path)
        return open_excel

    def get_sheet_data(self, index=None):
        '''
        加载所有sheet的内容,传下标
        '''
        sheet_name = self.load_excel().sheetnames
        if index == None:
            index = 1
        data = self.load_excel()[sheet_name[index]]
        return data

    def get_sheet_data_name(self, sheetname=None):
        '''
        加载所有sheet的内容,传Sheet_name
        '''
        wb = self.load_excel()
        try:
            if sheetname == None:
                sheetname = 'Sheet1'
            data = wb[sheetname]
            return data
        except:
            print('找不到对应的Sheet，请核实后输入')

    def get_cell_value(self, row, cols, index=None):
        '''
        获取某一个单元格内容
        '''
        data = self.get_sheet_data(index).cell(row=row, column=cols).value
        return data

    def get_rows(self, index=None):
        '''
        获取最大行数
        '''
        row = self.get_sheet_data(index).max_row
        return row

    def get_rows_value(self, row, index=None):
        '''
        获取某一行内容
        '''
        row_list = []
        for i in self.get_sheet_data(index)[row]:
            data = bool(re.search('^O[0-9]*', i.coordinate)) or bool(re.search('^P[0-9]*', i.coordinate))
            if data == False:
                row_list.append(i.value)
            else:
                row_list.append(None)
        return row_list

    def get_columns_value(self, index=None, cols=None):
        '''
        获取某一列得数据
        '''
        columns_list = []
        if cols == None:
            cols = 'A'
        columns_list_data = self.get_sheet_data(index)[cols]
        for i in columns_list_data:
            columns_list.append(i.value)
        return columns_list

    def excel_write_data(self, row, cols, value, sheet_name=None):
        '''
        写入数据
        '''
        wb = self.load_excel()
        if sheet_name == None:
            sheet_name = 'Common'
        wr = wb[sheet_name]
        wr.cell(row, cols, value)
        wb.save(dir_config.excel_path)

    def get_rows_number(self, case_id, index=None):
        '''
        获取行号
        '''
        num = 1
        cols_data = self.get_columns_value(index=index)
        for i in cols_data:
            if case_id == i:
                return num
            num = num + 1
        return num

    def get_excel_data(self, index=None):
        '''
        :param index: excel的下标
        :return: 获取excel里面所有的数据
        '''
        data_list = []
        for i in range(self.get_rows(index) - 1):
            data_list.append(self.get_rows_value(i + 2, index))
        return data_list


excel_data = HandExcel()
if __name__ == '__main__':
    # print(HandExcel().get_excel_data())
    # excel_data.excel_write_data(6, 1, 'hellotest')
    # print(excel_data.get_excel_data(9))
    print(excel_data.get_rows(9))
    test_data = excel_data.get_excel_data(9)
    print(test_data)
